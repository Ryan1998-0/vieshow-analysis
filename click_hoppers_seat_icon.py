import time
from datetime import datetime
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from config import THEATER_DETAIL_URL, RECORDS_FILE, THEATER_NAME


# 初始化 Excel 紀錄檔
def init_records():
    if not Path(RECORDS_FILE).exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "售票紀錄"
        headers = ["查詢時間", "影城", "電影", "日期", "場次", "語言", "session_id", "已售出", "總座位", "售出率%"]
        ws.append(headers)
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="2F5496")
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 8
        ws.column_dimensions["I"].width = 8
        ws.column_dimensions["J"].width = 10
        wb.save(RECORDS_FILE)


# 初始化 Chrome 瀏覽器驅動
def init_driver(headless=False):
    options = Options()
    if headless:
        options.add_argument('--headless=new')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
    service = Service()
    driver = webdriver.Chrome(service=service, options=options)
    return driver


# 開啟戲院頁面並點擊今天的日期頁籤
def click_today_date(driver):
    driver.get(THEATER_DETAIL_URL)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.TAG_NAME, 'body'))
    )

    today = datetime.now().strftime('%m/%d')
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, f"//a[contains(@class,'haveTime') and contains(text(), '{today}')]"))
    )
    date_tab = driver.find_element(By.XPATH, f"//a[contains(@class,'haveTime') and contains(text(), '{today}')]")
    date_tab.click()
    time.sleep(1)


# 取得指定電影今天的所有場次（可依版本篩選），回傳 (time_str, lang, session_id) 清單
def get_all_sessions(driver, movie_name, versions=None):
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, f"//h2[contains(text(), '{movie_name}')]"))
    )

    movie_day = driver.find_element(
        By.XPATH,
        f"//h2[contains(text(), '{movie_name}')]/following-sibling::div[contains(@class,'movieDay')][1]"
    )
    soup = BeautifulSoup(movie_day.get_attribute('innerHTML'), 'html.parser')

    sessions = []
    seen_langs = set()

    for h4 in soup.find_all('h4'):
        lang = h4.text.strip()

        # 同一版本名稱只取第一次出現，後續重複的屬於下一部電影
        if lang in seen_langs:
            continue
        seen_langs.add(lang)

        if versions and lang not in versions:
            continue

        ul = h4.find_next_sibling('ul', class_='bookList')
        if not ul:
            continue

        for li in ul.find_all('li'):
            time_a = li.find('a')
            if not time_a:
                continue
            time_str = time_a.text.strip()
            booking_href = time_a.get('href', '')
            if 'txtSessionId=' in booking_href:
                session_id = booking_href.split('txtSessionId=')[-1].split('&')[0]
                if time_str:
                    try:
                        h, m = map(int, time_str.split(':'))
                        session_dt = datetime.now().replace(hour=h, minute=m, second=0, microsecond=0)
                        if session_dt < datetime.now():
                            print(f'  略過已過場次：{time_str}')
                            continue
                    except ValueError:
                        pass
                    sessions.append((time_str, lang, session_id))

    return sessions


# 在戲院頁面點擊指定場次的座位圖示，切換至新分頁
def click_seat_icon(driver, session_id):
    seat_link = driver.find_element(
        By.XPATH,
        f"//a[contains(@href,'SessionSeats.aspx') and contains(@href,'txtSessionId={session_id}')]"
    )
    seat_link.click()
    time.sleep(3)

    if len(driver.window_handles) > 1:
        driver.switch_to.window(driver.window_handles[-1])

    WebDriverWait(driver, 30).until(
        EC.url_contains('SessionSeats.aspx')
    )


# 解析座位頁面，計算已售出與總座位數
def parse_seat_info(driver):
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.label-info, .label-danger'))
        )
    except Exception:
        pass

    soup = BeautifulSoup(driver.page_source, 'html.parser')
    sold = len(soup.find_all('div', class_='label-danger'))
    total = len(soup.find_all('div', class_=lambda c: c and ('label-danger' in c or 'label-info' in c)))
    return sold, total


# 將查詢結果寫入 Excel，若資料無變化則略過
def write_record(time_str, lang, session_id, sold, total, movie_title):
    init_records()
    wb = openpyxl.load_workbook(RECORDS_FILE)
    ws = wb.active
    date_str = datetime.now().strftime('%Y-%m-%d')
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    rate = round(sold / total * 100, 1) if total > 0 else 0

    for row in range(ws.max_row, 1, -1):
        existing_session = ws[f"G{row}"].value
        existing_date = ws[f"D{row}"].value
        existing_time = ws[f"E{row}"].value
        existing_sold = ws[f"H{row}"].value
        existing_total = ws[f"I{row}"].value
        if str(existing_session) == str(session_id) and existing_date == date_str and existing_time == time_str:
            if existing_sold == sold and existing_total == total:
                print(f'  [{time_str} {lang}] 無變化，略過')
                wb.close()
                return
            break

    ws.append([now, THEATER_NAME, movie_title, date_str, time_str, lang, session_id, sold, total, rate])
    wb.save(RECORDS_FILE)
